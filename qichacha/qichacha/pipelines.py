# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html

import json
import codecs

from openpyxl import Workbook
from openpyxl import load_workbook


class QichachaPipeline(object):
    def __init__(self):
        self.file = codecs.open('data.json', mode='wb', encoding='utf-8')#数据存储到data.json

        self.wb = load_workbook('test.xlsx') # 打开模版
        # self.wb = Workbook() # 新建xlsx
        self.ws = self.wb.active
        self.ws.append(['企业名称', '法定代表人', '注册资本', '成立时间', '邮箱', '电话', '地址'])  # 设置表头

    def parseT(self, value):
        return json.dumps(value).decode('unicode_escape')

    def process_item(self, item, spider):
        line = json.dumps(dict(item)) + "\n"
        self.file.write(line.decode("unicode_escape"))

        # test = [self.parseT(item['title'])]

        # 一些非正常数据, 空数据, 检测到的无效数据  需要过滤掉.
        test = ['11234']
        self.ws.append(test)
        self.wb.save('test.xlsx')

        return item